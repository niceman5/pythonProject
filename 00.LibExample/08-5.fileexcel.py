import openpyxl
import random
from pathlib import Path
from openpyxl.chart import Reference, Series, LineChart

root_dir = Path('C:\\')
file_dir = root_dir / '00.Dev' / 'study' / 'Python' / '00.LibExample' / 'file'
excel_file  = file_dir / 'sample.xlsx'

wb = openpyxl.load_workbook(excel_file)
# print(type(wb))
print(wb.get_sheet_names())  # sheet의 이름들을 출력

# ws = wb.get_sheet_by_name('Sheet1')  --> 이렇게 안씀.
# sheet에 접근
ws = wb['Sheet1']

# 엑셀의 각 셀에 접근
print(ws.max_column, ws.max_row, type(ws))
print(ws['A1'].value,ws['B1'].value,ws['C1'].value )
a3 = ws.cell(row=3, column=3)
print(a3.value)

print('A1 - > A2...순서대로 읽기')
for row in ws.rows:
    for cell in row:
        print(cell.value, ' ',  end='' )
    print('')

excel_file2 = file_dir / 'sample2.xlsx'
wb = openpyxl.Workbook()
# sheet지정 index->위치 title->sheet의 명
ws = wb.create_sheet(index=0, title='신규 sheet')
ws['a1'] = 100
wb.save(filename=excel_file2)

excel_file3 = file_dir / 'sample3_chart.xlsx'
wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0, title='챠트 sheet')
for i in range(10):
    ws.append([random.randint(1,10)])

values = Reference(ws, min_row=1, min_col=1, max_row=10, max_col=1)
series = Series(values, title='샘플 챠트')
chart = LineChart()
chart.append(series)
ws.add_chart(chart)
wb.save(filename=excel_file3)


